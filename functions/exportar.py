"""
Exportações Excel do GP Flow: Backlog, Curadoria e Sprint.
Todas usam o mesmo estilo visual (cabeçalho azul, zebra, bordas finas).
"""
import io
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.append(str(Path(__file__).parent.parent))
from functions.util import formatar_codigo, limpar_texto

FONTE = "Arial"
COR_CABECALHO = "1F4E78"
COR_SUBTOTAL = "DDEBF7"
COR_ZEBRA = "F2F2F2"

BORDA_FINA = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def _nova_planilha(titulo_aba: str, titulo_topo: str, colunas: list, larguras: list):
    wb = Workbook()
    ws = wb.active
    ws.title = titulo_aba

    ws["A1"] = titulo_topo
    ws["A1"].font = Font(name=FONTE, size=14, bold=True)

    linha_cab = 3
    for col, nome in enumerate(colunas, start=1):
        c = ws.cell(row=linha_cab, column=col, value=nome)
        c.font = Font(name=FONTE, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=COR_CABECALHO)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDA_FINA
        ws.column_dimensions[get_column_letter(col)].width = larguras[col - 1]
    ws.freeze_panes = f"A{linha_cab + 1}"
    return wb, ws, linha_cab


def _escrever_linhas(ws, linha_inicial: int, linhas: list) -> int:
    """Escreve uma lista de listas de valores a partir da linha_inicial, com zebra. Retorna a próxima linha livre."""
    linha = linha_inicial
    for zebra, valores in enumerate(linhas):
        for col, valor in enumerate(valores, start=1):
            c = ws.cell(row=linha, column=col, value=valor)
            c.font = Font(name=FONTE, size=10)
            c.border = BORDA_FINA
            if zebra % 2 == 1:
                c.fill = PatternFill("solid", start_color=COR_ZEBRA)
            c.alignment = Alignment(vertical="center", wrap_text=True)
        linha += 1
    return linha


def exportar_backlog_excel(df: pd.DataFrame) -> bytes:
    colunas = ["Id", "Título", "Solicitante", "Tipo", "Estado", "Prioridade", "Responsável", "Aging (dias)", "Score"]
    larguras = [10, 55, 22, 18, 16, 12, 22, 12, 10]
    wb, ws, linha_cab = _nova_planilha("Backlog", "Backlog de Demandas — GP Flow", colunas, larguras)

    linhas = [
        [
            int(row["id"]), row["titulo"], row.get("solicitante", ""),
            formatar_codigo(row.get("tipo")), row.get("estado", ""),
            formatar_codigo(row.get("prioridade")), row.get("responsavel_atendimento", ""),
            row.get("aging_dias", ""), row.get("score_exibicao", row.get("score", "")),
        ]
        for _, row in df.iterrows()
    ]
    _escrever_linhas(ws, linha_cab + 1, linhas)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def exportar_curadoria_excel(df: pd.DataFrame) -> bytes:
    colunas = ["Id", "Título", "Macroprocesso", "Sistema", "Score", "Observações",
               "Solicitante", "Responsável", "Data criação", "Aging (dias)"]
    larguras = [10, 50, 18, 18, 10, 40, 22, 22, 16, 12]
    wb, ws, linha_cab = _nova_planilha("Curadoria", "Curadoria — GP Flow", colunas, larguras)

    linhas = [
        [
            int(row["id"]), row["titulo"], row.get("macroprocesso", "") or "",
            row.get("sistema", "") or "", row.get("score", ""), row.get("observacoes", "") or "",
            row.get("solicitante", ""), row.get("responsavel_atendimento", ""),
            row.get("data_criacao", ""), row.get("aging_dias", ""),
        ]
        for _, row in df.iterrows()
    ]
    _escrever_linhas(ws, linha_cab + 1, linhas)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _adicionar_aba_homeoffice(wb, data_inicio, data_fim, responsaveis):
    """
    Adiciona a aba 'Planejamento HomeOffice': grade em branco (responsáveis x
    dias úteis da sprint) para preencher no Excel, com a legenda de cores.
    """
    ws = wb.create_sheet("Planejamento HomeOffice")

    # gera os dias úteis (seg-sex) entre início e fim da sprint
    dias = []
    if data_inicio and data_fim:
        try:
            ini = pd.to_datetime(data_inicio).date()
            fim = pd.to_datetime(data_fim).date()
            d = ini
            while d <= fim:
                if d.weekday() < 5:  # 0-4 = seg-sex
                    dias.append(d)
                d += pd.Timedelta(days=1)
        except Exception:
            dias = []

    # separa em semanas (blocos de 5 dias úteis) como no modelo
    dias_pt = {0: "segunda-feira", 1: "terça-feira", 2: "quarta-feira",
               3: "quinta-feira", 4: "sexta-feira"}
    meses_pt = {1: "janeiro", 2: "fevereiro", 3: "março", 4: "abril", 5: "maio",
                6: "junho", 7: "julho", 8: "agosto", 9: "setembro", 10: "outubro",
                11: "novembro", 12: "dezembro"}

    def rotulo(d):
        return f"{dias_pt[d.weekday()]}, {d.day} de {meses_pt[d.month]} de {d.year}"

    fundo_cab = PatternFill("solid", fgColor="D9D9D9")
    fundo_nome = PatternFill("solid", fgColor="D9D9D9")
    fonte_cab = Font(name=FONTE, size=10, bold=True)
    fonte_norm = Font(name=FONTE, size=10)
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

    semanas = [dias[i:i + 5] for i in range(0, len(dias), 5)]
    linha = 1

    # dropdown com as opções da legenda (validação de dados do Excel)
    opcoes = "FÉRIAS,HOMEOFFICE/SEDE,HOMEOFFICE,FERIADO,HU,TORRE,SEDE,ATESTADO,BANCO"
    dv = DataValidation(type="list", formula1=f'"{opcoes}"', allow_blank=True)
    dv.error = "Escolha uma opção da lista."
    dv.errorTitle = "Valor inválido"
    dv.prompt = "Selecione o local/situação do dia."
    dv.promptTitle = "Planejamento"
    ws.add_data_validation(dv)

    for semana in semanas:
        # linha de cabeçalho: "Sprint" + as datas
        c = ws.cell(row=linha, column=1, value="Sprint")
        c.font = fonte_cab; c.fill = fundo_cab; c.alignment = centro; c.border = BORDA_FINA
        for j, d in enumerate(semana):
            c = ws.cell(row=linha, column=2 + j, value=rotulo(d))
            c.font = fonte_cab; c.fill = fundo_cab; c.alignment = centro; c.border = BORDA_FINA
        # uma linha por responsável (célula com dropdown para preencher)
        for k, nome in enumerate(responsaveis):
            r = linha + 1 + k
            c = ws.cell(row=r, column=1, value=nome)
            c.font = fonte_cab; c.fill = fundo_nome; c.alignment = centro; c.border = BORDA_FINA
            for j in range(len(semana)):
                c = ws.cell(row=r, column=2 + j)
                c.font = fonte_norm; c.alignment = centro; c.border = BORDA_FINA
                dv.add(c)  # aplica o dropdown nesta célula
        linha = linha + 1 + len(responsaveis) + 1  # +1 linha em branco entre semanas

    # legenda de cores
    legenda = [
        ("FÉRIAS", "C55A11", "FFFFFF"),
        ("HOMEOFFICE/SEDE", "FFA500", "000000"),
        ("HOMEOFFICE", "FFFF00", "000000"),
        ("FERIADO", "00B050", "FFFFFF"),
        ("HU", "FFFFFF", "000000"),
        ("TORRE", "C6E0B4", "000000"),
        ("SEDE", "00B0F0", "000000"),
        ("ATESTADO", "4472C4", "FFFFFF"),
        ("BANCO", "0070C0", "FFFFFF"),
    ]
    linha += 1
    for nome, cor_fundo, cor_texto in legenda:
        c = ws.cell(row=linha, column=2, value=nome)
        c.fill = PatternFill("solid", fgColor=cor_fundo)
        c.font = Font(name=FONTE, size=10, bold=True, color=cor_texto)
        c.alignment = centro
        c.border = BORDA_FINA
        linha += 1

    # larguras
    ws.column_dimensions["A"].width = 12
    for j in range(len(dias[:5]) or [1]):
        ws.column_dimensions[get_column_letter(2 + j)].width = 26


def exportar_sprint_excel(nome_sprint: str, df_demandas: pd.DataFrame, df_atividades: pd.DataFrame = None,
                          data_inicio=None, data_fim=None, responsaveis=None) -> bytes:
    colunas = ["Id", "Título", "Tipo de entrada", "Status", "Responsável", "Impedimento"]
    larguras = [10, 55, 16, 16, 20, 35]
    wb, ws, linha_cab = _nova_planilha("Sprint", nome_sprint, colunas, larguras)

    linhas = [
        [
            int(row["id"]), row["titulo"], row.get("tipo_entrada", ""),
            row.get("status_kanban", ""), row.get("responsavel_sprint", "") or "",
            limpar_texto(row.get("impedimento")),
        ]
        for _, row in df_demandas.iterrows()
    ]
    if df_atividades is not None:
        linhas += [
            [
                "", row["titulo"], row.get("tipo_entrada", ""),
                row.get("status_kanban", ""), row.get("responsavel_sprint", "") or "",
                limpar_texto(row.get("impedimento")),
            ]
            for _, row in df_atividades.iterrows()
        ]
    _escrever_linhas(ws, linha_cab + 1, linhas)

    # segunda aba: planejamento de home office (grade em branco)
    if responsaveis:
        _adicionar_aba_homeoffice(wb, data_inicio, data_fim, responsaveis)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
